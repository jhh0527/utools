from __future__ import annotations

from pathlib import Path

from utube.categories import category_label
from utube.format_util import format_published
from utube.models import VideoItem

_EXCEL_HEADERS = (
    "순위",
    "제목",
    "번역",
    "채널",
    "카테고리",
    "조회수",
    "좋아요",
    "댓글",
    "업로드",
    "길이",
    "쇼츠",
    "URL",
)


def export_videos_excel(
    path: Path,
    rows: list[VideoItem],
    *,
    title_ko: dict[str, str] | None = None,
) -> None:
    """``rows`` 를 ``.xlsx`` 로 저장. ``openpyxl`` 필요."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError("엑셀 저장에 openpyxl 이 필요합니다: pip install openpyxl") from e

    from utube.translate_util import is_mostly_korean, translate_to_korean

    ko_map = title_ko or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Utube"

    header_font = Font(bold=True)
    for col, title in enumerate(_EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, v in enumerate(rows, start=2):
        if is_mostly_korean(v.title):
            translated = ""
        else:
            translated = ko_map.get(v.video_id) or translate_to_korean(v.title)
        ws.cell(row=i, column=1, value=i - 1)
        ws.cell(row=i, column=2, value=v.title)
        ws.cell(row=i, column=3, value=translated)
        ws.cell(row=i, column=4, value=v.channel)
        ws.cell(row=i, column=5, value=category_label(v.category_id))
        ws.cell(row=i, column=6, value=v.view_count)
        ws.cell(row=i, column=7, value=v.like_count if v.like_count is not None else "")
        ws.cell(row=i, column=8, value=v.comment_count if v.comment_count is not None else "")
        ws.cell(row=i, column=9, value=format_published(v.published_at))
        ws.cell(row=i, column=10, value=v.duration)
        ws.cell(row=i, column=11, value=v.shorts_display)
        link = ws.cell(row=i, column=12, value=v.url)
        link.hyperlink = v.url
        link.style = "Hyperlink"

    widths = (6, 48, 48, 22, 14, 12, 10, 10, 12, 10, 8, 52)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{max(1, len(rows) + 1)}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
